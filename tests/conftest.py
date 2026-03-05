import pytest
import os
from pathlib import Path
from unittest.mock import MagicMock
import json

@pytest.fixture
def mock_litellm(monkeypatch):
    """LiteLLMのcompletionをモック化するフィクスチャ"""
    mock = MagicMock()
    
    # デフォルトのレスポンス
    mock_response = MagicMock()
    mock_response.choices = [
        MagicMock(message=MagicMock(content=json.dumps({"test": "data"})))
    ]
    mock.return_value = mock_response
    
    monkeypatch.setattr("litellm.completion", mock)
    return mock

@pytest.fixture
def sample_excel_path(tmp_path):
    """テスト用の空のExcelファイルを生成するフィクスチャ"""
    excel_path = tmp_path / "test.xlsx"
    # 実際の内容は必要ない場合が多いが、extractorがopenpyxlを使うので
    # 最低限のファイルを作成するか、extractor自体をモック化する
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "テストデータ"
    ws["B1"] = "報告者"
    wb.save(excel_path)
    return excel_path

@pytest.fixture
def output_dir(tmp_path):
    """テスト用の出力ディレクトリ"""
    d = tmp_path / "output"
    d.mkdir()
    return d

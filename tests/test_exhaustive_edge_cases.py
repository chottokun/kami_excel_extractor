import pytest
import openpyxl
from pathlib import Path
from kami_excel_extractor.extractor import MetadataExtractor
from kami_excel_extractor.core import KamiExcelExtractor

@pytest.fixture
def exhaustive_excel(tmp_path):
    wb = openpyxl.Workbook()
    
    # 1. 巨大なシート (100x100)
    ws_large = wb.active
    ws_large.title = "LargeSheet"
    for r in range(1, 101):
        for c in range(1, 27):
            ws_large.cell(row=r, column=c, value=f"Val_{r}_{c}")
            
    # 2. 壊れた計算式を含むシート
    ws_broken = wb.create_sheet("BrokenLogic")
    ws_broken["A1"] = "#REF!"
    ws_broken["A2"] = "=SUM(UNKNOWN_REF)"
    
    # 3. 多数の画像 (10枚)
    ws_images = wb.create_sheet("ManyImages")
    from PIL import Image as PILImage
    img_path = tmp_path / "small.png"
    PILImage.new('RGB', (10, 10), color='blue').save(img_path)
    from openpyxl.drawing.image import Image as OpenpyxlImage
    for i in range(1, 11):
        img = OpenpyxlImage(str(img_path))
        img.anchor = f"A{i}"
        ws_images.add_image(img)
        
    # 4. 隠しシート
    ws_hidden = wb.create_sheet("HiddenSheet")
    ws_hidden["A1"] = "Secret Data"
    ws_hidden.sheet_state = 'hidden'
    
    excel_path = tmp_path / "exhaustive_test.xlsx"
    wb.save(excel_path)
    return excel_path

def test_extractor_handling_large_sheet(tmp_path, exhaustive_excel):
    """巨大なシートの抽出が時間内・正常に終わるか検証"""
    me = MetadataExtractor(output_dir=tmp_path)
    result = me.extract(exhaustive_excel)
    assert "LargeSheet" in result["sheets"]
    assert len(result["sheets"]["LargeSheet"]["cells"]) >= 2600

def test_extractor_handling_broken_logic(tmp_path, exhaustive_excel):
    """壊れた計算式があってもクラッシュしないか検証"""
    me = MetadataExtractor(output_dir=tmp_path)
    result = me.extract(exhaustive_excel, include_logic=True)
    cells = result["sheets"]["BrokenLogic"]["cells"]
    # クラッシュせず、formula が取得できている（または適切に処理されている）こと
    assert any(c["coord"] == "A2" for c in cells)

def test_extractor_handling_many_images(tmp_path, exhaustive_excel):
    """多数の画像がある場合に、すべてのアンカーが抽出されるか検証"""
    me = MetadataExtractor(output_dir=tmp_path)
    result = me.extract(exhaustive_excel)
    media = result["sheets"]["ManyImages"]["media"]
    assert len(media) == 10
    coords = [m["coord"] for m in media]
    assert "A1" in coords
    assert "A10" in coords

def test_extractor_handling_hidden_sheets(tmp_path, exhaustive_excel):
    """隠しシートも正しく抽出対象に含まれるか検証"""
    me = MetadataExtractor(output_dir=tmp_path)
    result = me.extract(exhaustive_excel)
    assert "HiddenSheet" in result["sheets"]
    assert "Secret Data" in result["sheets"]["HiddenSheet"]["html"]

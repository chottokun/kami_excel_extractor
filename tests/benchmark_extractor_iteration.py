import time
from pathlib import Path
import openpyxl
from kami_excel_extractor.extractor import MetadataExtractor

def create_large_excel(path, rows=1000, cols=20):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=f"Data {r}-{c}")

    # Add many different styles to test caching
    for r in range(1, 100):
        for c in range(1, cols + 1):
             ws.cell(row=r, column=c).font = openpyxl.styles.Font(bold=(r % 2 == 0))
             if r % 3 == 0:
                 ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Add some merged cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    wb.save(path)

def main():
    test_file = Path("benchmark_large.xlsx")
    create_large_excel(test_file, rows=1000, cols=50)
    extractor = MetadataExtractor(output_dir="temp_media")
    print("Starting benchmark...")
    start_time = time.time()
    for _ in range(5):
        extractor.extract(test_file, include_logic=False)
    end_time = time.time()
    duration = end_time - start_time
    print(f"Benchmark duration: {duration:.4f} seconds")
    if test_file.exists():
        test_file.unlink()

if __name__ == "__main__":
    main()

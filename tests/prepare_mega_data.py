
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from pathlib import Path
import io
from PIL import Image

def create_mega_test_excel(output_path: Path):
    wb = openpyxl.Workbook()
    
    # Sheet 1: 画像地獄
    ws1 = wb.active
    ws1.title = "ImageHeaven"
    ws1["A1"] = "複数画像とネストされた名前のテスト"
    
    img = Image.new('RGB', (100, 50), color=(255, 0, 0))
    img_path = Path("temp_mega.png")
    img.save(img_path)
    
    from openpyxl.drawing.image import Image as OpenpyxlImage
    for i in range(5):
        oi = OpenpyxlImage(str(img_path))
        ws1.add_image(oi, f"B{i*5 + 2}")
    
    # Sheet 2: 巨大な空白とBounding Box
    ws2 = wb.create_sheet("SparseData")
    ws2["A1"] = "Start"
    ws2["Z100"] = "End" # 遠くに一つだけデータを置く
    ws2.cell(row=50, column=10, value="Middle Data")
    
    # Sheet 3: 数式と結合
    ws3 = wb.create_sheet("LogicAndMerged")
    ws3.merge_cells("A1:C3")
    ws3["A1"] = "Merged Header"
    ws3["A4"] = 100
    ws3["B4"] = 200
    ws3["C4"] = "=A4+B4"

    wb.save(output_path)
    img_path.unlink()
    print(f"Created Mega Sample: {output_path}")

if __name__ == "__main__":
    create_mega_test_excel(Path("tests/assets/mega_mixed_report.xlsx"))

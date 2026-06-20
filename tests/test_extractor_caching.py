import openpyxl
from pathlib import Path
from kami_excel_extractor.extractor import MetadataExtractor
from openpyxl.styles import PatternFill, Font, Border, Side

def test_metadata_extractor_caching(tmp_path):
    # Create a test Excel file with varying styles
    xlsx_path = tmp_path / "test_cache.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Cell A1: Red background, bold
    ws["A1"].value = "Header"
    ws["A1"].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    ws["A1"].font = Font(bold=True)

    # Cell A2: Same style as A1
    ws["A2"].value = "Header 2"
    ws["A2"].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    ws["A2"].font = Font(bold=True)

    # Cell B1: Blue background
    ws["B1"].value = "Data"
    ws["B1"].fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")

    # Cell B2: Same style as B1
    ws["B2"].value = "Data 2"
    ws["B2"].fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")

    wb.save(xlsx_path)

    extractor = MetadataExtractor(output_dir=tmp_path / "output")
    result = extractor.extract(xlsx_path)

    sheet = result["sheets"]["Sheet"]
    cells = {c["coord"]: c for c in sheet["cells"]}

    # Verify values and styles are correct
    assert cells["A1"]["value"] == "Header"
    assert cells["A1"]["style"]["bold"] is True
    assert "background-color: #FF0000" in sheet["html"]

    assert cells["A2"]["value"] == "Header 2"
    assert cells["A2"]["style"]["bold"] is True

    assert cells["B1"]["value"] == "Data"
    assert "background-color: #0000FF" in sheet["html"]

    assert cells["B2"]["value"] == "Data 2"

    # Verify that style_id-based caching worked correctly
    # We can check the internal cache after extraction
    assert len(extractor._style_cache) >= 2

    # In openpyxl, A1 and A2 should have the same style_id
    wb_reload = openpyxl.load_workbook(xlsx_path)
    ws_reload = wb_reload.active
    assert ws_reload["A1"].style_id == ws_reload["A2"].style_id
    assert ws_reload["B1"].style_id == ws_reload["B2"].style_id
    assert ws_reload["A1"].style_id != ws_reload["B1"].style_id

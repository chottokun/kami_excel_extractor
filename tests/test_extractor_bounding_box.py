import pytest
import openpyxl
from unittest.mock import MagicMock
from kami_excel_extractor.extractor import MetadataExtractor

@pytest.fixture
def extractor(tmp_path):
    return MetadataExtractor(output_dir=tmp_path)

def test_get_bounding_box_empty(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    # An empty sheet in openpyxl might have max_row=1, max_column=1 by default if not manipulated
    # but ws.max_row/max_column represent the extent of cells that have been accessed or have values.

    # In _get_bounding_box:
    # min_r, max_r = 1, 0
    # if ws.max_row > 0: ...
    # return 1, max_r, 1, max_c

    bounds = extractor._get_bounding_box(ws)
    assert bounds == (1, 0, 1, 0)

def test_get_bounding_box_with_values(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=3).value = "Hello"
    ws.cell(row=5, column=1).value = "World"

    bounds = extractor._get_bounding_box(ws)
    # max_r should be 5, max_c should be 3
    assert bounds == (1, 5, 1, 3)

def test_get_bounding_box_with_merged_cells(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells(range_string="B2:D4")

    bounds = extractor._get_bounding_box(ws)
    # max_r should be 4, max_c should be 4
    assert bounds == (1, 4, 1, 4)

def test_get_bounding_box_with_values_and_merged_cells(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=2).value = "Inside"
    ws.merge_cells(range_string="A1:E5")

    bounds = extractor._get_bounding_box(ws)
    assert bounds == (1, 5, 1, 5)

def test_get_bounding_box_with_images_from(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Mock an image with anchor._from
    mock_img = MagicMock()
    mock_img.anchor._from.row = 10 # 0-indexed, so row 11
    mock_img.anchor._from.col = 5  # 0-indexed, so col 6

    ws._images = [mock_img]

    bounds = extractor._get_bounding_box(ws)
    # max_r = max(0, 10+1) = 11
    # max_c = max(0, 5+1) = 6
    assert bounds == (1, 11, 1, 6)

def test_get_bounding_box_with_images_no_from(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Mock an image WITHOUT anchor._from
    mock_img = MagicMock()
    del mock_img.anchor._from

    ws._images = [mock_img]

    bounds = extractor._get_bounding_box(ws)
    assert bounds == (1, 0, 1, 0)

def test_get_bounding_box_complex(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=3, column=3).value = "Value" # max_r=3, max_c=3
    ws.merge_cells(range_string="E1:F2")     # max_r=3, max_c=6

    mock_img = MagicMock()
    mock_img.anchor._from.row = 7 # row 8
    mock_img.anchor._from.col = 1 # col 2
    ws._images = [mock_img]

    bounds = extractor._get_bounding_box(ws)
    # max_r = max(3, 2, 8) = 8
    # max_c = max(3, 6, 2) = 6
    assert bounds == (1, 8, 1, 6)

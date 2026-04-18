import pytest
from pathlib import Path
from kami_excel_extractor.extractor import MetadataExtractor
from kami_excel_extractor.core import KamiExcelExtractor
from kami_excel_extractor.schema import ExtractionOptions
import openpyxl

@pytest.fixture
def logic_excel(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LogicTest"
    
    # 結合セルを追加してシンプルテーブル判定を回避
    ws.merge_cells("A1:C1")
    ws["A1"] = "計算テスト報告書"
    
    # 1. 計算式と数値
    ws["A3"] = 100
    ws["B3"] = 200
    ws["C3"] = "=A3+B3" # 計算式
    
    # 2. 表示形式（単位）
    ws["A4"] = 1500
    ws["A4"].number_format = '"¥"#,##0' # JPY
    
    ws["B4"] = 0.5
    ws["B4"].number_format = '0.00%' # PERCENT
    
    excel_path = tmp_path / "logic_test.xlsx"
    wb.save(excel_path)
    return excel_path

def test_extractor_extracts_logic_when_enabled(tmp_path, logic_excel):
    """include_logic=True の場合に計算式と単位が抽出されることを検証"""
    extractor = MetadataExtractor(output_dir=tmp_path)
    
    # ON の場合
    result_on = extractor.extract(logic_excel, include_logic=True)
    html_on = result_on["sheets"]["LogicTest"]["html"]
    
    assert 'data-formula="=A3+B3"' in html_on
    assert 'data-unit="JPY"' in html_on
    assert 'data-unit="PERCENT"' in html_on
    
    # メタデータ内
    cells = result_on["sheets"]["LogicTest"]["cells"]
    c3_meta = next(c for c in cells if c["coord"] == "C3")
    assert c3_meta["formula"] == "=A3+B3"
    
    # OFF の場合
    result_off = extractor.extract(logic_excel, include_logic=False)
    html_off = result_off["sheets"]["LogicTest"]["html"]
    
    assert 'data-formula' not in html_off
    # unitは現状常に抽出される仕様（パフォーマンス影響が少ないため）だが、
    # もし制限したい場合はここでアサーションを追加
    
@pytest.mark.asyncio
async def test_core_passes_logic_to_prompt(tmp_path, logic_excel):
    """KamiExcelExtractor が include_logic をプロンプトに反映させているか検証"""
    from unittest.mock import patch, MagicMock
    
    extractor = KamiExcelExtractor(output_dir=tmp_path)
    
    with patch("litellm.acompletion") as mock_completion:
        mock_completion.return_value.choices[0].message.content = '{"data": "ok"}'
        
        options = ExtractionOptions(include_logic=True, use_visual_context=False)
        await extractor.aextract_structured_data(logic_excel, options=options)
        
        # 呼ばれた際のメッセージを確認
        args, kwargs = mock_completion.call_args
        messages = kwargs["messages"]
        user_msg = next(m["content"] for m in messages if m["role"] == "user")
        text_content = next(c["text"] for c in user_msg if c["type"] == "text" and "データソース" in c["text"])
        
        # 指示が含まれているか
        assert "data-formula属性" in text_content
        assert "data-unit属性" in text_content
        # HTML内に計算式が含まれているか
        assert 'data-formula="=A3+B3"' in text_content

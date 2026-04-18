import pytest
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from kami_excel_extractor.extractor import MetadataExtractor

@pytest.fixture
def enhanced_excel(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StyleTest"
    
    # 1. 太い罫線付きセル
    thick_side = Side(style='thick', color='000000')
    ws["A1"] = "Thick Border"
    ws["A1"].border = Border(top=thick_side, bottom=thick_side)
    
    # 2. 結合セル + 背景色 + 太字
    ws.merge_cells("B2:C2")
    ws["B2"] = "Rich Cell"
    ws["B2"].font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    ws["B2"].fill = yellow_fill
    
    excel_path = tmp_path / "style_test.xlsx"
    wb.save(excel_path)
    return excel_path

def test_enhanced_extraction(tmp_path, enhanced_excel):
    extractor = MetadataExtractor(output_dir=tmp_path)
    result = extractor.extract(enhanced_excel)
    
    sheet_data = result["sheets"]["StyleTest"]
    html_content = sheet_data["html"]
    
    # A. スタイル情報（罫線）の確認
    # thick border -> 3px solid black 程度に変換されるはず
    assert 'border-top: 3px solid black' in html_content
    assert 'border-bottom: 3px solid black' in html_content
    
    # B. 背景色とフォントスタイルの確認
    assert 'background-color: #FFFF00' in html_content
    assert 'font-weight: bold' in html_content
    
    # C. 座標データ（data-coord）の確認
    assert 'data-coord="A1"' in html_content
    assert 'data-coord="B2"' in html_content
    
    # D. cells メタデータの確認
    cells = sheet_data["cells"]
    assert len(cells) > 0
    a1_meta = next(c for c in cells if c["coord"] == "A1")
    assert a1_meta["style"]["borders"]["top"] == "thick"
    
    b2_meta = next(c for c in cells if c["coord"] == "B2")
    assert b2_meta["colspan"] == 2
    assert b2_meta["style"]["bold"] is True

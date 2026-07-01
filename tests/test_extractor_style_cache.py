from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
from openpyxl.styles import Border, Font, PatternFill, Side

from src.kami_excel_extractor.extractor import MetadataExtractor


def test_style_cache_same_style_id(tmp_path):
    """同一のstyle_idを持つ複数のセルが正しくキャッシュされ、結果が同一になることをテストする。"""
    extractor = MetadataExtractor(tmp_path)

    # モックのセルオブジェクトを作成
    cell1 = MagicMock()
    cell1.coordinate = "A1"
    cell1.style_id = 42
    cell1.number_format = "¥#,##0"
    cell1.value = 1000

    cell2 = MagicMock()
    cell2.coordinate = "B1"
    cell2.style_id = 42
    cell2.number_format = "¥#,##0"
    cell2.value = 2000

    # 共通のスタイル属性を設定
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    font = Font(bold=True)

    for cell in (cell1, cell2):
        cell.fill = fill
        cell.border = border
        cell.font = font

    # 最初のセルで各情報を取得し、キャッシュを構築させる
    unit1 = extractor._get_unit_info(cell1)
    border_info1 = extractor._get_border_info(cell1)
    css1 = extractor._get_cell_style_string(cell1)

    # キャッシュに保存されていることを確認
    assert 42 in extractor._style_cache
    assert 42 in extractor._unit_cache
    assert 42 in extractor._border_cache

    # cell2の属性アクセス時に例外を発生させるようにし、キャッシュが優先されることを確認する。
    cell2.fill = MagicMock(side_effect=AssertionError("Should not access cell properties if cached"))
    cell2.border = MagicMock(side_effect=AssertionError("Should not access cell properties if cached"))
    cell2.font = MagicMock(side_effect=AssertionError("Should not access cell properties if cached"))

    # cell2に対して呼び出しても、例外が発生せずに同一のキャッシュ結果が返ることを確認する。
    assert extractor._get_unit_info(cell2) == unit1
    assert extractor._get_border_info(cell2) == border_info1
    assert extractor._get_cell_style_string(cell2) == css1


def test_style_cache_missing_style_id(tmp_path):
    """style_id属性を持たないセルオブジェクトが渡された場合でも、キャッシュせず安全に動作することをテストする。"""
    extractor = MetadataExtractor(tmp_path)

    cell = MagicMock(spec=[])
    # style_idを定義しない
    assert not hasattr(cell, "style_id")

    cell.number_format = "General"
    cell.value = "test"
    cell.fill = None
    cell.border = None
    cell.font = None

    # 例外なく動作することを確認
    unit = extractor._get_unit_info(cell)
    border_info = extractor._get_border_info(cell)
    css = extractor._get_cell_style_string(cell)

    assert unit is None
    assert border_info == {}
    assert css == ""

    # キャッシュには何も入らないことを確認
    assert len(extractor._style_cache) == 0
    assert len(extractor._unit_cache) == 0
    assert len(extractor._border_cache) == 0


def test_bold_cache_integration(tmp_path):
    """MetadataExtractorのデータ抽出ループで _bold_cache が正しく使われ、太字判定がキャッシュされることをテストする。"""
    extractor = MetadataExtractor(tmp_path)

    # 実際の一時Excelファイルを作成
    excel_path = tmp_path / "test_bold.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Test1"
    ws["A2"] = "Test2"

    # 太字のフォントを設定（同じスタイルが適用される）
    font = Font(bold=True)
    ws["A1"].font = font
    ws["A2"].font = font

    wb.save(excel_path)
    wb.close()

    # extractを実行（_generate_metadata_and_html が内部で呼ばれる）
    wb_loaded = openpyxl.load_workbook(excel_path, data_only=True)
    ws_loaded = wb_loaded.active

    html_tbl, cell_metadata = extractor._generate_metadata_and_html(ws_loaded)

    # A1 と A2 の style_id が同一であることを確認し、キャッシュに保存されていることを確認
    style_id_a1 = getattr(ws_loaded["A1"], "style_id", None)
    style_id_a2 = getattr(ws_loaded["A2"], "style_id", None)

    assert style_id_a1 == style_id_a2
    assert style_id_a1 in extractor._bold_cache
    assert extractor._bold_cache[style_id_a1] is True
    wb_loaded.close()


def test_datetime_and_date_preservation(tmp_path):
    """日付 (date) や日時 (datetime) オブジェクトが HTML 上で isoformat に変換され、デグレードしていないことをテストする。"""
    from datetime import date, datetime

    extractor = MetadataExtractor(tmp_path)

    excel_path = tmp_path / "test_date.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    test_date = date(2026, 7, 1)
    test_datetime = datetime(2026, 7, 1, 12, 30, 45)

    ws["A1"] = test_date
    ws["B1"] = test_datetime

    wb.save(excel_path)
    wb.close()

    wb_loaded = openpyxl.load_workbook(excel_path, data_only=True)
    ws_loaded = wb_loaded.active

    html_tbl, cell_metadata = extractor._generate_metadata_and_html(ws_loaded)

    # HTML上に ISO 形式文字列が出現することを確認
    assert "2026-07-01" in html_tbl
    assert "2026-07-01T12:30:45" in html_tbl
    wb_loaded.close()

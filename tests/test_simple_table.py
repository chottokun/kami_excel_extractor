import pytest
import openpyxl
from datetime import date, datetime
from pathlib import Path
from kami_excel_extractor.extractor import MetadataExtractor

@pytest.fixture
def extractor(tmp_path):
    return MetadataExtractor(output_dir=tmp_path)

def test_is_simple_table_valid(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.append(["Value1", "Value2"])
    assert extractor.is_simple_table(ws) is True

def test_is_simple_table_merged_cells(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.merge_cells("A1:B1")
    assert extractor.is_simple_table(ws) is False

def test_is_simple_table_one_header(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Header1"])
    ws.append(["Value1"])
    assert extractor.is_simple_table(ws) is False

def test_is_simple_table_one_row(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    assert extractor.is_simple_table(ws) is False

def test_extract_simple_table_basic(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])

    data = extractor.extract_simple_table(ws)
    assert data == [
        {"Name": "Alice", "Age": 30},
        {"Name": "Bob", "Age": 25}
    ]

def test_extract_simple_table_dates(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Event", "Date"])
    d = date(2023, 1, 1)
    dt = datetime(2023, 1, 1, 12, 0, 0)
    ws.append(["Start", d])
    ws.append(["Exact", dt])

    data = extractor.extract_simple_table(ws)
    assert data == [
        {"Event": "Start", "Date": "2023-01-01"},
        {"Event": "Exact", "Date": "2023-01-01T12:00:00"}
    ]

def test_extract_simple_table_missing_headers(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1: One header, one empty
    ws.cell(row=1, column=1, value="ID")
    # Row 2: Data
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="NoHeader")

    data = extractor.extract_simple_table(ws)
    assert data == [
        {"ID": 1, "Column2": "NoHeader"}
    ]

def test_extract_simple_table_empty_rows(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Name"])
    ws.append([1, "Alice"])
    ws.append([None, None]) # Empty row
    ws.append([2, "Bob"])

    data = extractor.extract_simple_table(ws)
    assert data == [
        {"ID": 1, "Name": "Alice"},
        {"ID": 2, "Name": "Bob"}
    ]

def test_extract_simple_table_none_values(extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Note"])
    ws.append([1, None])

    data = extractor.extract_simple_table(ws)
    # 現在の実装では None の値は辞書に含まれない
    assert data == [
        {"ID": 1}
    ]

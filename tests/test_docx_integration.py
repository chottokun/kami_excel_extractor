from pathlib import Path

import pytest

from kami_excel_extractor.core import KamiExcelExtractor
from kami_excel_extractor.schema import RagOptions


@pytest.mark.asyncio
async def test_integration_docx_generation(tmp_path):
    # ダミーのExcelファイルは tests/assets 以下の既存のものか、新規に作るか。
    # 既存のテスト資産 tests/prepare_test_data.py 等を参考にする。
    # ここでは既存の assets があるか、もしくは簡単な Excel ファイルを生成するか。
    # 実データ complex_report.xlsx がルートにあるので、それを使用。
    excel_path = Path("complex_report.xlsx")
    if not excel_path.exists():
        # もし存在しなければ、モックまたは別のものでテストする。
        # 既存の conftest.py や test_core.py でどうしているか参考にする。
        pytest.skip("complex_report.xlsx is not found at project root")

    extractor = KamiExcelExtractor(output_dir=tmp_path)

    opts = RagOptions(
        model="gemini/gemini-1.5-flash",  # テスト時はLLMの呼び出しが発生する可能性があるが、
        # モックまたはキャッシュが動作することを期待するか、あるいはLLMをダミーにする。
        # 既存のテスト（test_core_rag_chunks.pyなど）はどのようにLLMを呼んでいるか。
        # 通常、テスト環境では API_KEY がない場合があり、モックされているか、
        # もしくは is_simple が True で LLM をバイパスするシートだけでテストする。
        # ここでは is_simple なシートを含むダミーExcelを作成してテストするのが最も確実。
        use_visual_context=False,
        include_visual_summaries=False,
    )

    # is_simple のダミーを作成
    import openpyxl

    dummy_excel = tmp_path / "dummy_simple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SimpleSheet"
    ws.append(["Name", "Age"])
    ws.append(["Alice", "30"])
    wb.save(dummy_excel)

    # docx の生成を直接呼び出し
    docx_path, structured_data = await extractor.aextract_docx(dummy_excel, options=opts)

    assert docx_path.exists()
    assert docx_path.suffix == ".docx"
    assert "SimpleSheet" in structured_data["sheets"]


@pytest.mark.asyncio
async def test_cli_docx_option(tmp_path):
    import argparse

    import openpyxl

    from kami_excel_extractor.cli import run_async

    dummy_excel = tmp_path / "dummy_simple_cli.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SimpleSheet"
    ws.append(["Name", "Age"])
    ws.append(["Bob", "40"])
    wb.save(dummy_excel)

    args = argparse.Namespace(
        input=str(dummy_excel),
        model=None,
        api_key=None,
        base_url=None,
        timeout=600.0,
        rpm=None,
        output_dir=str(tmp_path),
        no_vision=True,
        rag=False,
        rag_format="yaml_frontmatter",
        max_chunk_chars=1000,
        chunk_overlap_lines=2,
        no_coordinates=False,
        no_logic_annotations=False,
        system_prompt=None,
        dpi=150,
        include_logic=False,
        visual_summaries=False,
        verbose=True,
        docx=True,
    )

    await run_async(args)

    expected_docx = tmp_path / "dummy_simple_cli.docx"
    assert expected_docx.exists()


@pytest.mark.asyncio
async def test_cli_rag_docx_format(tmp_path):
    import argparse
    import json

    import openpyxl

    from kami_excel_extractor.cli import run_async

    dummy_excel = tmp_path / "dummy_simple_rag.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SimpleSheet"
    ws.append(["Name", "Age"])
    ws.append(["Charlie", "50"])
    wb.save(dummy_excel)

    args = argparse.Namespace(
        input=str(dummy_excel),
        model=None,
        api_key=None,
        base_url=None,
        timeout=600.0,
        rpm=None,
        output_dir=str(tmp_path),
        no_vision=True,
        rag=True,
        rag_format="docx",
        max_chunk_chars=1000,
        chunk_overlap_lines=2,
        no_coordinates=False,
        no_logic_annotations=False,
        system_prompt=None,
        dpi=150,
        include_logic=False,
        visual_summaries=False,
        verbose=True,
        docx=False,
    )

    await run_async(args)

    expected_docx = tmp_path / "dummy_simple_rag.docx"
    assert expected_docx.exists()

    # RAGメタデータJSONファイルも確認
    expected_rag_json = tmp_path / "dummy_simple_rag_rag.json"
    assert expected_rag_json.exists()

    with open(expected_rag_json, "r", encoding="utf-8") as f:
        meta_data = json.load(f)
    assert "docx_path" in meta_data
    assert meta_data["docx_path"].endswith("dummy_simple_rag.docx")

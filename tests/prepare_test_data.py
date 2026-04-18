import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from pathlib import Path

def create_complex_kami_excel(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ComplexLayout"

    # スタイル定義
    thick = Side(style='thick', color='000000')
    thin = Side(style='thin', color='000000')
    header_fill = PatternFill(start_color="CCE5FF", fill_type="solid") # 水色
    section_fill = PatternFill(start_color="E0E0E0", fill_type="solid") # グレー

    # 1. ヘッダーエリア (結合セル)
    ws.merge_cells("A1:F2")
    ws["A1"] = "現場調査報告書 (2026年度)"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = header_fill
    for row in ws["A1:F2"]:
        for cell in row:
            cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)

    # 2. 基本情報セクション (ラベルと値がバラバラ)
    ws["A4"] = "調査日:"
    ws["B4"] = "2026-04-18"
    ws["D4"] = "担当者:"
    ws["E4"] = "田中 太郎"
    
    # 罫線で囲む
    for coord in ["A4", "B4", "D4", "E4"]:
        ws[coord].border = Border(bottom=thin)

    # 3. 複雑な入れ子構造のテーブル
    # 見出し
    ws.merge_cells("A6:A8")
    ws["A6"] = "カテゴリ"
    ws.merge_cells("B6:C6")
    ws["B6"] = "点検項目"
    ws["B7"] = "内容"
    ws["C7"] = "基準"
    ws.merge_cells("D6:F6")
    ws["D6"] = "判定・備考"
    
    for coord in ["A6", "B6", "B7", "C7", "D6"]:
        ws[coord].fill = section_fill
        ws[coord].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # データ行1
    ws.merge_cells("A9:A10")
    ws["A9"] = "構造体"
    ws["B9"] = "ひび割れ"
    ws["C9"] = "0.3mm以下"
    ws["D9"] = "適合"
    ws.merge_cells("E9:F10")
    ws["E9"] = "特記事項なし"
    
    # データ行2
    ws["B10"] = "鉄筋露出"
    ws["C10"] = "なし"
    ws["D10"] = "適合"

    # 全体に罫線を適用
    for row in ws["A6:F10"]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    wb.save(path)

if __name__ == "__main__":
    create_complex_kami_excel(Path("tests/assets/complex_kami_sample.xlsx"))

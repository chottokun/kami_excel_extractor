import pytest
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
from kami_excel_extractor.extractor import MetadataExtractor

@pytest.fixture
def sample_html_excel(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HTML_Test"
    
    # 1. 普通のセル
    ws["A1"] = "Normal Text"
    
    # 2. 結合セル (B2:C3 -> colspan=2, rowspan=2)
    ws.merge_cells("B2:C3")
    ws["B2"] = "Merged Cell"
    
    # 3. 背景色付きセル
    yellow_fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    ws["A3"].fill = yellow_fill
    ws["A3"] = "Colored Cell"

    # 4. 改行を含むセル
    ws["A4"] = "Line 1\nLine 2"
    
    excel_path = tmp_path / "html_test.xlsx"
    wb.save(excel_path)
    return excel_path

def test_extract_html_table(tmp_path, sample_html_excel):
    extractor = MetadataExtractor(output_dir=tmp_path)
    result = extractor.extract(sample_html_excel)
    
    assert "sheets" in result
    assert "HTML_Test" in result["sheets"]
    
    html_content = result["sheets"]["HTML_Test"]["html"]
    
    # テーブル構造が保たれているか
    assert "<table border='1'>" in html_content
    assert "</table>" in html_content
    
    # 1. 普通のセル
    assert "<td>Normal Text</td>" in html_content
    
    # 2. 結合セル (colspan=2, rowspan=2が展開され、左上以外はskipされていること)
    assert 'colspan="2"' in html_content
    assert 'rowspan="2"' in html_content
    assert ">Merged Cell</td>" in html_content
    
    # 3. 背景色付きセル (00FFFF00 -> FFFF00 に変換される)
    assert 'style="background-color: #FFFF00"' in html_content
    assert ">Colored Cell</td>" in html_content

    # 4. 改行を含むセル (\n が <br> に置換されていること)
    assert "<td>Line 1<br>Line 2</td>" in html_content

import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch
import openpyxl
from kami_excel_extractor.extractor import MetadataExtractor

@pytest.fixture
def extractor(tmp_path):
    return MetadataExtractor(output_dir=tmp_path)

@pytest.fixture
def sample_excel_file(tmp_path):
    wb = openpyxl.Workbook()
    # Sheet 1: Simple table
    ws1 = wb.active
    ws1.title = "SimpleSheet"
    ws1.append(["ID", "Name"])
    ws1.append([1, "Alice"])
    ws1.append([2, "Bob"])

    # Sheet 2: Complex table (with merged cells)
    ws2 = wb.create_sheet("ComplexSheet")
    ws2.append(["Header 1", "Header 2"])
    ws2.merge_cells("A1:B1")
    ws2["A2"] = "Data"

    excel_path = tmp_path / "test.xlsx"
    wb.save(excel_path)
    return excel_path

def test_extract_orchestration(extractor):
    """Test the orchestration logic of extract() by mocking internal methods."""
    mock_wb = MagicMock()
    mock_wb.sheetnames = ["Sheet1", "Sheet2"]

    mock_ws1 = MagicMock()
    mock_ws2 = MagicMock()
    mock_wb.__getitem__.side_effect = lambda name: mock_ws1 if name == "Sheet1" else mock_ws2

    with patch("openpyxl.load_workbook", return_value=mock_wb) as mock_load_wb, \
         patch.object(MetadataExtractor, "_generate_metadata_and_html") as mock_gen_both, \
         patch.object(MetadataExtractor, "_extract_media") as mock_ext_media, \
         patch.object(MetadataExtractor, "is_simple_table") as mock_is_simple, \
         patch.object(MetadataExtractor, "extract_simple_table") as mock_ext_simple:

        # Configure mocks
        mock_gen_both.return_value = ("<table></table>", [])
        mock_ext_media.return_value = [{"coord": "A1", "filename": "img.png"}]
        mock_is_simple.side_effect = [True, False]
        mock_ext_simple.return_value = [{"ID": 1, "Name": "Alice"}]

        dummy_path = Path("dummy.xlsx")
        result = extractor.extract(dummy_path)

        # Verify load_workbook call
        mock_load_wb.assert_called_once_with(dummy_path, data_only=True)

        # Verify calls
        assert mock_gen_both.call_count == 2
        assert mock_ext_media.call_count == 2
        assert mock_is_simple.call_count == 2
        mock_ext_simple.assert_called_once_with(mock_ws1)

        # Verify result structure
        assert "sheets" in result
        assert "Sheet1" in result["sheets"]
        assert "Sheet2" in result["sheets"]

        # Sheet 1 (Simple)
        s1 = result["sheets"]["Sheet1"]
        assert s1["is_simple"] is True
        assert s1["structured_data"] == [{"ID": 1, "Name": "Alice"}]
        assert s1["html"] == "<table></table>"
        assert s1["media"] == [{"coord": "A1", "filename": "img.png"}]

        # Sheet 2 (Complex)
        s2 = result["sheets"]["Sheet2"]
        assert s2["is_simple"] is False
        assert "structured_data" not in s2
        assert s2["html"] == "<table></table>"

def test_extract_integration_minimal(extractor, sample_excel_file):
    """Test extract() with a real file to ensure all components work together."""
    result = extractor.extract(sample_excel_file)

    assert "sheets" in result
    assert "SimpleSheet" in result["sheets"]
    assert "ComplexSheet" in result["sheets"]

    # SimpleSheet verification
    ss = result["sheets"]["SimpleSheet"]
    assert ss["is_simple"] is True
    assert len(ss["structured_data"]) == 2
    assert ss["structured_data"][0]["ID"] == 1
    assert "<table" in ss["html"]

    # ComplexSheet verification
    cs = result["sheets"]["ComplexSheet"]
    assert cs["is_simple"] is False
    assert "structured_data" not in cs
    assert "<table" in cs["html"]

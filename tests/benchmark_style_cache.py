import time
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from kami_excel_extractor.extractor import MetadataExtractor

def create_styled_excel(path, rows=1000, cols=50):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create a few styles
    fill1 = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    fill2 = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    font1 = Font(bold=True)
    border1 = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value=f"Data {r}-{c}")
            if r % 2 == 0:
                cell.fill = fill1
                cell.font = font1
            else:
                cell.fill = fill2
                cell.border = border1

    wb.save(path)

def main():
    test_file = Path("benchmark_styled.xlsx")
    if not test_file.exists():
        print("Creating styled Excel file...")
        create_styled_excel(test_file, rows=1000, cols=50) # 50,000 cells

    extractor = MetadataExtractor(output_dir="temp_media")

    print("Starting benchmark...")
    # Warm up
    extractor.extract(test_file, include_logic=True)

    start_time = time.time()
    iterations = 3
    for i in range(iterations):
        print(f"Iteration {i+1}/{iterations}...")
        extractor.extract(test_file, include_logic=True)
    end_time = time.time()

    duration = end_time - start_time
    print(f"Total duration for {iterations} iterations: {duration:.4f} seconds")
    print(f"Average duration: {duration / iterations:.4f} seconds")

    # Cleanup if you want, but maybe keep for next run
    # if test_file.exists():
    #     test_file.unlink()

if __name__ == "__main__":
    main()

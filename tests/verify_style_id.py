import openpyxl
from openpyxl.styles import PatternFill
import os

def test_style_id():
    wb = openpyxl.Workbook()
    ws = wb.active

    fill1 = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    fill2 = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

    ws["A1"].fill = fill1
    ws["A2"].fill = fill1
    ws["B1"].fill = fill2

    print(f"A1 style_id: {ws['A1'].style_id}")
    print(f"A2 style_id: {ws['A2'].style_id}")
    print(f"B1 style_id: {ws['B1'].style_id}")

    assert ws["A1"].style_id == ws["A2"].style_id
    assert ws["A1"].style_id != ws["B1"].style_id
    print("Verification successful: style_id is consistent for identical styles.")

if __name__ == "__main__":
    test_style_id()

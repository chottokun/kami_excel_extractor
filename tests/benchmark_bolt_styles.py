import time
from pathlib import Path
import openpyxl
import os
import sys

# Add src to sys.path
sys.path.append(str(Path(__file__).parent.parent / "src"))

from kami_excel_extractor.extractor import MetadataExtractor

def benchmark():
    test_file = Path("benchmark_v2.xlsx")
    rows, cols = 1000, 50 # 50,000 cells
    wb = openpyxl.Workbook()
    ws = wb.active

    # Setup some styles
    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = openpyxl.styles.Font(bold=True)
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    print(f"Creating test file with {rows}x{cols} cells...")
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value=f"Data {r}-{c}")
            if r % 2 == 0:
                cell.fill = yellow_fill
            if c % 5 == 0:
                cell.font = bold_font
            if r % 3 == 0:
                cell.border = thin_border
            if c % 10 == 0:
                cell.number_format = '#,##0'

    wb.save(test_file)
    print("Test file created.")

    extractor = MetadataExtractor(output_dir="temp_media")

    n_iters = 5
    print(f"Benchmarking extraction {n_iters} times...")
    durations = []
    for i in range(n_iters):
        start_time = time.time()
        extractor.extract(test_file, include_logic=False)
        duration = time.time() - start_time
        durations.append(duration)
        print(f"Iteration {i+1}: {duration:.4f}s")

    avg_duration = sum(durations) / n_iters
    print(f"Average duration: {avg_duration:.4f} seconds")

    if test_file.exists():
        test_file.unlink()
    if Path("temp_media").exists():
        import shutil
        shutil.rmtree("temp_media")

if __name__ == "__main__":
    benchmark()

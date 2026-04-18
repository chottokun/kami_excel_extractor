import pytest
from kami_excel_extractor.utils import clean_kami_text
from kami_excel_extractor.extractor import MetadataExtractor
from pathlib import Path
import openpyxl

def test_clean_kami_text_logic():
    """クリーニング関数の単体テスト"""
    # 漢字の間の空白
    assert clean_kami_text("氏　　名") == "氏名"
    assert clean_kami_text("氏  名") == "氏名"
    
    # ひらがな・カタカナの間の空白
    assert clean_kami_text("あ　い　う") == "あいう"
    assert clean_kami_text("カ　タ　カ　ナ") == "カタカナ"
    
    # 文末・文頭の空白
    assert clean_kami_text("  氏名  ") == "氏名"
    
    # 英単語間の空白は維持（1つ）
    assert clean_kami_text("Hello World") == "Hello World"
    # 英単語間の不自然な多重空白は維持（現在の仕様）
    # ※英単語を壊さないように、日本語文字クラスのみを結合対象としている
    assert clean_kami_text("A   B   C") == "A   B   C"

def test_extractor_applies_cleaning(tmp_path):
    """Extractorが抽出時にクリーニングを適用しているか検証"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "担　当　者"
    ws["B1"] = "東　京　支　店"
    
    excel_path = tmp_path / "clean_test.xlsx"
    wb.save(excel_path)
    
    extractor = MetadataExtractor(output_dir=tmp_path)
    result = extractor.extract(excel_path)
    
    html = result["sheets"]["Sheet"]["html"]
    # HTML内でクリーニングされていること
    assert "担当者" in html
    assert "東京支店" in html
    assert "担　当　者" not in html
    
    # メタデータ内でクリーニングされていること
    cells = result["sheets"]["Sheet"]["cells"]
    a1_val = next(c["value"] for c in cells if c["coord"] == "A1")
    assert a1_val == "担当者"

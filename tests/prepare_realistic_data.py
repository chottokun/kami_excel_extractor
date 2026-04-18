import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from pathlib import Path
import io
from PIL import Image as PILImage

def create_business_report_excel(path: Path):
    wb = openpyxl.Workbook()
    
    # シート1: 概要とグラフ
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "月次売上報告書"
    ws1["A1"].font = Font(size=14, bold=True)
    
    # ダミーの表
    headers = ["月", "目標", "実績", "達成率"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h).fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    
    data = [
        ["2024-01", 1000, 1100, "110%"],
        ["2024-02", 1000, 950, "95%"],
        ["2024-03", 1200, 1300, "108%"]
    ]
    for r, row_data in enumerate(data, 4):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
    
    # グラフ領域のシミュレーション（画像を配置）
    img_path = path.parent / "temp_chart.png"
    PILImage.new('RGB', (200, 100), color='blue').save(img_path)
    from openpyxl.drawing.image import Image as OpenpyxlImage
    img = OpenpyxlImage(str(img_path))
    img.anchor = "E3"
    ws1.add_image(img)
    
    # シート2: 現場写真と詳細
    ws2 = wb.create_sheet("SitePhotos")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "現場状況写真"
    ws2["A1"].alignment = Alignment(horizontal="center")
    
    # 写真枠 (A3, C3)
    temp_files = []
    for i, coord in enumerate(["A3", "C3"]):
        ws2[coord] = f"(写真添付エリア {coord})"
        ws2[coord].border = Border(top=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'), bottom=Side(style='medium'))
        
        # 個別の画像ファイルを作成
        p_path = path.parent / f"temp_photo_{i}.png"
        PILImage.new('RGB', (100, 100), color=('green' if i==0 else 'yellow')).save(p_path)
        temp_files.append(p_path)
        
        photo = OpenpyxlImage(str(p_path))
        photo.anchor = coord
        ws2.add_image(photo)
        
    ws2["A15"] = "所見:"
    ws2.merge_cells("A16:D20")
    ws2["A16"] = "全体の構造に異常なし。経年劣化による微細なひび割れが認められるが、即時の補修は不要と判断される。"
    ws2["A16"].alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)
    
    # 保存後に削除
    if img_path.exists(): img_path.unlink()
    for f in temp_files:
        if f.exists(): f.unlink()

if __name__ == "__main__":
    assets_dir = Path("tests/assets")
    assets_dir.mkdir(parents=True, exist_ok=True)
    create_business_report_excel(assets_dir / "realistic_business_report.xlsx")
    print(f"Created: {assets_dir / 'realistic_business_report.xlsx'}")

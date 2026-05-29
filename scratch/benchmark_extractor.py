import time
from pathlib import Path
import sys
import os

# Add src to sys.path
sys.path.append(os.path.abspath("src"))

from kami_excel_extractor.extractor import MetadataExtractor
import openpyxl

def create_large_excel(path, rows=1000, cols=20):
    wb = openpyxl.Workbook()
    ws = wb.active
    from openpyxl.styles import PatternFill, Border, Side, Font

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font = Font(bold=True)

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value=f"Row {r} Col {c}")
            if r % 2 == 0:
                cell.fill = fill
                cell.border = border
                cell.font = font

    wb.save(path)

test_xlsx = Path("scratch/test_large.xlsx")
if not test_xlsx.exists():
    print("Creating large test Excel...")
    create_large_excel(test_xlsx, rows=2000, cols=30)

extractor = MetadataExtractor(output_dir="scratch/output")

print("Starting extraction...")
start_time = time.time()
result = extractor.extract(test_xlsx)
end_time = time.time()

print(f"Extraction took {end_time - start_time:.4f} seconds")

# Run again to see if it's consistent
start_time = time.time()
result = extractor.extract(test_xlsx)
end_time = time.time()
print(f"Second run took {end_time - start_time:.4f} seconds")

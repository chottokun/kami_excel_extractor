import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font

# Add src to sys.path
sys.path.append(os.path.abspath("src"))

from kami_excel_extractor.extractor import MetadataExtractor

def test_cache_logic():
    path = Path("scratch/test_cache.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create two cells with same style
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    font = Font(bold=True)

    ws["A1"].value = "Cell 1"
    ws["A1"].fill = fill
    ws["A1"].font = font

    ws["B1"].value = "Cell 2"
    ws["B1"].fill = fill
    ws["B1"].font = font

    wb.save(path)

    extractor = MetadataExtractor(output_dir="scratch/output_verify")

    # Patch _get_border_info to count calls
    original_get_border_info = extractor._get_border_info
    call_count = 0
    def mocked_get_border_info(cell):
        nonlocal call_count
        call_count += 1
        return original_get_border_info(cell)

    extractor._get_border_info = mocked_get_border_info

    print("Extracting...")
    result = extractor.extract(path)

    # A1 and B1 have same style. Bounding box might be small.
    # Total cells in bounding box (1,1, 1,2) is 2.
    # If caching works, call_count should be 1 (for s_id of A1 and B1).
    # Wait, openpyxl might have multiple style_ids even if they look same?
    # Let's check style_ids.
    wb_loaded = openpyxl.load_workbook(path)
    ws_loaded = wb_loaded.active
    print(f"A1 style_id: {ws_loaded['A1'].style_id}")
    print(f"B1 style_id: {ws_loaded['B1'].style_id}")

    print(f"Total _get_border_info calls: {call_count}")

    # Verification
    sheet_data = result["sheets"]["Sheet"]
    html = sheet_data["html"]
    cells = sheet_data["cells"]

    assert "background-color: #FFFF00" in html
    assert "font-weight: bold" in html
    assert len(cells) == 2
    assert cells[0]["value"] == "Cell 1"
    assert cells[1]["value"] == "Cell 2"
    assert cells[0]["style"]["bold"] is True
    assert cells[1]["style"]["bold"] is True

    if ws_loaded['A1'].style_id == ws_loaded['B1'].style_id:
        if call_count == 1:
            print("SUCCESS: Style cache utilized!")
        else:
            print(f"FAILURE: Style cache NOT utilized? Calls: {call_count}")
    else:
        print(f"INFO: Style IDs differ ({ws_loaded['A1'].style_id} vs {ws_loaded['B1'].style_id}), cache won't trigger for these.")

if __name__ == "__main__":
    test_cache_logic()

import openpyxl
import json
import os
from pathlib import Path
from openpyxl.utils import get_column_letter
import io

def get_border_info(cell):
    borders = {}
    if cell.border:
        if cell.border.left.style: borders["L"] = cell.border.left.style
        if cell.border.right.style: borders["R"] = cell.border.right.style
        if cell.border.top.style: borders["T"] = cell.border.top.style
        if cell.border.bottom.style: borders["B"] = cell.border.bottom.style
    return borders

def extract_media(ws, output_media_dir, sheet_name):
    """ワークシートから画像を抽出し、座標情報を返す"""
    media_info = []
    if not hasattr(ws, "_images"):
        return media_info

    for idx, img in enumerate(ws._images):
        row = img.anchor._from.row + 1
        col = img.anchor._from.col + 1
        coord = f"{get_column_letter(col)}{row}"
        
        image_filename = f"{sheet_name}_img_{coord}_{idx}.png"
        save_path = output_media_dir / image_filename
        
        try:
            # openpyxlのImageオブジェクトからバイナリを取得
            # refが直接BytesIOやファイルオブジェクトを指している場合がある
            if hasattr(img.ref, "read"):
                data = img.ref.read()
            elif hasattr(img.ref, "getvalue"):
                data = img.ref.getvalue()
            else:
                # 最終手段としてプロパティを確認
                data = img.ref
            
            with open(save_path, "wb") as f:
                f.write(data)
            
            media_info.append({
                "coord": coord,
                "filename": str(image_filename),
                "type": "image"
            })
        except Exception as e:
            print(f"Failed to extract image at {coord}: {e}")
            
    return media_info

def extract_comprehensive_map(filename, output_dir):
    wb = openpyxl.load_workbook(filename, data_only=True)
    output_media_dir = Path(output_dir) / "media"
    output_media_dir.mkdir(parents=True, exist_ok=True)
    
    full_map = { "sheets": {} }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        merged_cells_map = {}
        for m_range in ws.merged_cells.ranges:
            m_range_str = str(m_range)
            for r in range(m_range.min_row, m_range.max_row + 1):
                for c in range(m_range.min_col, m_range.max_col + 1):
                    coord = f"{get_column_letter(c)}{r}"
                    merged_cells_map[coord] = m_range_str
        
        cells_map = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                has_style = cell.fill and cell.fill.start_color.index != '00000000'
                has_border = bool(get_border_info(cell))
                if cell.value is None and not has_style and not has_border: continue
                cell_info = {"c": cell.coordinate, "v": str(cell.value) if cell.value is not None else ""}
                if has_style: cell_info["bg"] = str(cell.fill.start_color.index)
                if has_border: cell_info["b"] = get_border_info(cell)

                if cell.coordinate in merged_cells_map:
                    cell_info["m"] = merged_cells_map[cell.coordinate]

                cells_map.append(cell_info)
        
        media_list = extract_media(ws, output_media_dir, sheet_name)
        full_map["sheets"][sheet_name] = {
            "cells": cells_map,
            "media": media_list
        }
    return full_map

if __name__ == "__main__":
    res = extract_comprehensive_map("complex_report.xlsx", "data/output")
    with open("data/output/comprehensive_map.json", "w", encoding="utf-8") as f:
        json.dump(res, f, ensure_ascii=False, indent=2)
    print(f"Extracted data and media to data/output/")

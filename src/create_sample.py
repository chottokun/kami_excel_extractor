import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

def create_sample_excel(filename="sample_hoganshi.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "報告書"

    # 方眼紙のようにセルを細かく設定（列幅を狭く）
    for i in range(1, 21):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 3

    # スタイル設定
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_header = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    font_bold = Font(bold=True)

    # タイトル
    ws.merge_cells("B2:H3")
    ws["B2"] = "業務完了報告書"
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].font = Font(size=14, bold=True)

    # 項目1: 報告者（方眼紙形式の結合）
    ws.merge_cells("B5:C5")
    ws["B5"] = "報告者"
    ws["B5"].fill = fill_header
    ws["B5"].border = border_thin
    
    ws.merge_cells("D5:G5") # 入力欄
    ws["D5"] = "山田 太郎"
    ws["D5"].border = border_thin

    # 項目2: 日付
    ws.merge_cells("B6:C6")
    ws["B6"] = "日付"
    ws["B6"].fill = fill_header
    ws["B6"].border = border_thin
    
    ws.merge_cells("D6:G6")
    ws["D6"] = "2026-03-02"
    ws["D6"].border = border_thin

    # 項目3: 内容（大きな結合セル）
    ws.merge_cells("B8:B12")
    ws["B8"] = "報告内容"
    ws["B8"].fill = fill_header
    ws["B8"].border = border_thin
    ws["B8"].alignment = Alignment(vertical="center")

    ws.merge_cells("C8:H12")
    ws["C8"] = """本日、プロジェクトAの主要なマイルストーンを達成しました。
詳細は以下の通りです。
- XMLマップ抽出ロジックの構築
- Gemini 2.0 Flashによる構造化テスト"""
    ws["C8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["C8"].border = border_thin

    # 明細テーブル（標準的な表形式も混在）
    ws["B14"] = "作業明細"
    ws["B14"].font = font_bold

    headers = ["ID", "作業名", "時間(h)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=15, column=i+2)
        cell.value = h
        cell.fill = fill_header
        cell.border = border_thin

    data = [
        [1, "環境構築", 2.0],
        [2, "コード実装", 4.5],
        [3, "テスト", 1.5]
    ]
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=16+r_idx, column=c_idx+2)
            cell.value = val
            cell.border = border_thin

    wb.save(filename)
    print(f"Created: {filename}")

if __name__ == "__main__":
    create_sample_excel()

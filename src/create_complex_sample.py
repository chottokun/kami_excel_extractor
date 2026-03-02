import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
from PIL import Image as PILImage, ImageDraw
import io
import os

def create_dummy_image(text, filename="dummy.png", color="red"):
    img = PILImage.new('RGB', (400, 300), color=(200, 200, 200))
    draw = ImageDraw.Draw(img)
    draw.text((150, 140), text, fill="black")
    draw.rectangle([50, 50, 350, 250], outline=color, width=5)
    img.save(filename)
    return filename

def create_complex_excel(filename="complex_report.xlsx"):
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: 基本報告 (方眼紙形式) ---
    ws1 = wb.active
    ws1.title = "報告概要"
    for i in range(1, 31):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 2
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_header = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    ws1.merge_cells("B2:L3")
    ws1["B2"] = "施工不良箇所 調査報告書"
    ws1["B2"].font = Font(size=16, bold=True)
    ws1["B2"].alignment = Alignment(horizontal="center")

    # 方眼紙特有の細かいラベル配置
    ws1.merge_cells("B5:D5")
    ws1["B5"] = "調査日"
    ws1["B5"].fill = fill_header
    ws1["B5"].border = border_thin
    
    ws1.merge_cells("E5:H5")
    ws1["E5"] = "2026/03/02"
    ws1["E5"].border = border_thin

    ws1.merge_cells("B7:D15")
    ws1["B7"] = "特記事項"
    ws1["B7"].fill = fill_header
    ws1["B7"].border = border_thin
    ws1["B7"].alignment = Alignment(vertical="center", horizontal="center")

    ws1.merge_cells("E7:R15")
    ws1["E7"] = "北側外壁の3階付近に、設計図にないクラックを確認。座標 A-15付近の写真を参照のこと。"
    ws1["E7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws1["E7"].border = border_thin

    # --- Sheet 2: 写真資料 (画像混在) ---
    ws2 = wb.create_sheet("現場写真")
    ws2["A1"] = "写真 No.1: クラック発生状況"
    ws2["A1"].font = Font(bold=True)
    
    # ダミー画像の作成と配置
    img_path = create_dummy_image("Crack at North Wall", "crack.png")
    img = Image(img_path)
    ws2.add_image(img, "A3") # A3セルを起点に配置

    ws2["A20"] = "写真 No.2: 全体俯瞰"
    img_path2 = create_dummy_image("General View", "general.png", color="blue")
    img2 = Image(img_path2)
    ws2.add_image(img2, "A22")

    wb.save(filename)
    # クリーンアップ
    os.remove("crack.png")
    os.remove("general.png")
    print(f"Created complex sample: {filename}")

if __name__ == "__main__":
    create_complex_excel()

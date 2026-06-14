"""
Excelファイルから詳細なメタデータ、視覚スタイル、およびロジックを抽出するエンジン。
"""

import html
import io
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from PIL import Image

from .utils import clean_kami_text, secure_filename

logger = logging.getLogger(__name__)

# セキュリティ設定: 画像の最大ピクセル数と最大バイト数
MAX_IMAGE_PIXELS = 25000000  # 25MP
MAX_IMAGE_BYTES = 20971520  # 20MB
Image.MAX_IMAGE_PIXELS = MAX_IMAGE_PIXELS


class MetadataExtractor:
    """
    Excelワークブックからテキスト、構造、スタイル、および埋め込みメディアを抽出する高精度エクストラクター。

    Attributes:
        output_dir (Path): 抽出されたメディア（画像）を保存するディレクトリ。
        media_dir (Path): メディアファイルの具体的な保存先。
    """

    def __init__(self, output_dir: Union[str, Path]):
        """
        MetadataExtractorを初期化する。

        Args:
            output_dir: 解析結果（主に画像）を出力するディレクトリパス。
        """
        self.output_dir = Path(output_dir)
        self.media_dir = self.output_dir / "media"
        self.media_dir.mkdir(parents=True, exist_ok=True)
        self._style_cache = {}  # style_id -> (style_str, unit, borders, bold)

    def _get_border_info(self, cell: openpyxl.cell.Cell) -> Dict[str, str]:
        """
        セルの罫線情報を取得し、各辺のスタイルを辞書形式で返す。

        Args:
            cell: openpyxlのセルオブジェクト。

        Returns:
            Dict[str, str]: {'left': 'thin', 'top': 'thick', ...} 形式の辞書。
        """
        borders = {}
        if cell.border:
            sides = ["left", "right", "top", "bottom"]
            for side in sides:
                border_side = getattr(cell.border, side)
                if border_side and border_side.style:
                    borders[side] = border_side.style
        return borders

    def _get_cell_style_string(self, cell: openpyxl.cell.Cell) -> str:
        """
        セルの視覚的属性（色、線、フォント）をCSS形式の文字列に変換する。

        LLMがテーブルの構造（ヘッダー、セクションの区切り）を理解するための重要なヒントとなる。

        Args:
            cell: openpyxlのセルオブジェクト。

        Returns:
            str: "background-color: #FFFFFF; border-top: 1px solid black;" 等のCSS文字列。
        """
        styles = []

        # 背景色の抽出 (ARGBをRGBに変換)
        if cell.fill and hasattr(cell.fill, "start_color") and cell.fill.start_color:
            c_idx = str(cell.fill.start_color.index)
            if c_idx not in ("00000000", "0"):
                if len(c_idx) == 8:  # ARGB (Alpha-RGB)
                    c_idx = c_idx[2:]
                if all(c in "0123456789ABCDEFabcdef" for c in c_idx):
                    styles.append(f"background-color: #{c_idx}")

        # 罫線情報をCSSのborder属性に近似
        border_info = self._get_border_info(cell)
        style_map = {
            "medium": "2px solid",
            "thick": "3px solid",
            "thin": "1px solid",
            "dashed": "1px dashed",
            "dotted": "1px dotted",
        }

        for side, style in border_info.items():
            css_style = style_map.get(style, "1px solid")
            styles.append(f"border-{side}: {css_style} black")

        # フォントウェイト
        if cell.font:
            if cell.font.b:
                styles.append("font-weight: bold")
            if cell.font.i:
                styles.append("font-style: italic")

        return "; ".join(styles)

    def _get_unit_info(self, cell: openpyxl.cell.Cell) -> Optional[str]:
        """
        セルの表示形式(Number Format)からデータの単位や型を推測する。

        Args:
            cell: openpyxlのセルオブジェクト。

        Returns:
            Optional[str]: 'JPY', 'PERCENT', 'DATE' 等の識別子。
        """
        fmt = cell.number_format
        if not fmt or fmt == "General":
            return None

        fmt_lower = fmt.lower()
        if "¥" in fmt or "jpy" in fmt_lower:
            return "JPY"
        if "$" in fmt:
            return "USD"
        if "%" in fmt:
            return "PERCENT"
        if "yy" in fmt_lower or "mm" in fmt_lower or "dd" in fmt_lower:
            return "DATE"
        return fmt

    def _parse_image_anchor(self, anchor: Any) -> Tuple[Optional[int], Optional[int]]:
        """
        画像のアンカー情報から行番号と列番号（1-indexed）をパースする。

        Args:
            anchor: openpyxlの画像アンカーオブジェクト。

        Returns:
            Tuple[Optional[int], Optional[int]]: (行番号, 列番号)。取得できない場合はNone。
        """
        row, col = None, None
        if hasattr(anchor, "_from"):  # TwoCellAnchor
            row = anchor._from.row + 1
            col = anchor._from.col + 1
        elif hasattr(anchor, "row"):  # OneCellAnchor
            row = anchor.row + 1
            col = anchor.col + 1
        elif isinstance(anchor, str):  # String anchor (e.g. "A1")
            row, col = coordinate_to_tuple(anchor)
        return row, col

    def _get_image_raw_data(self, img: Any, coord: str, sheet_name: str) -> Optional[bytes]:
        """
        画像リファレンスからRAWバイナリデータを取得し、サイズ制限を適用する。

        Args:
            img: openpyxlの画像オブジェクト。
            coord: 画像の位置座標文字列。
            sheet_name: ワークシート名。

        Returns:
            Optional[bytes]: 抽出された画像データ。スキップまたはエラー時はNone。
        """
        raw_data = None
        # 1. テスト用 Mock の特別なハンドリング (無限ループ防止)
        if "Mock" in type(img.ref).__name__:
            raw_data = img.ref.read() if hasattr(img.ref, "read") else img.ref.getvalue()
        # 2. すでにバイト列として存在する場合 (メモリコピーを最小化)
        elif isinstance(img.ref, (bytes, bytearray, memoryview)):
            raw_data = img.ref
        # 3. バッファに直接アクセス可能な場合 (最も効率的: BytesIO等)
        elif hasattr(img.ref, "getbuffer"):
            raw_data = img.ref.getbuffer()
        # 4. ストリーム(readメソッド)を持つ場合、チャンクごとに読み込む
        elif hasattr(img.ref, "read"):
            raw_data_buf = io.BytesIO()
            total_read = 0
            chunk_size = 8192
            while True:
                chunk = img.ref.read(chunk_size)
                if not chunk:
                    break
                total_read += len(chunk)
                if total_read > MAX_IMAGE_BYTES:
                    logger.warning(f"Skipping large image at {coord} on {sheet_name} (stream exceeds limit)")
                    return None
                raw_data_buf.write(chunk)

            if total_read <= MAX_IMAGE_BYTES:
                raw_data = raw_data_buf.getbuffer()
        # 5. その他の getvalue フォールバック
        elif hasattr(img.ref, "getvalue"):
            raw_data = img.ref.getvalue()
        else:
            raise AttributeError("Image reference has no readable data attribute")

        # 共通のサイズチェック
        if raw_data is not None and len(raw_data) > MAX_IMAGE_BYTES:
            logger.warning(f"Skipping large image at {coord} on {sheet_name} (size: {len(raw_data)} bytes)")
            return None

        return raw_data

    def _process_single_image(self, img: Any, idx: int, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        単一の画像を処理し、保存してメタデータを返す。

        Args:
            img: openpyxlの画像オブジェクト。
            idx: シート内での画像のインデックス。
            sheet_name: シート名。

        Returns:
            Optional[Dict[str, Any]]: 画像のメタデータ。スキップされた場合はNone。
        """
        try:
            row, col = self._parse_image_anchor(img.anchor)
        except Exception as e:
            logger.warning(f"Failed to parse coordinate anchor '{img.anchor}' on sheet {sheet_name}: {e}")
            row, col = None, None

        coord = f"{get_column_letter(col)}{row}" if (isinstance(row, int) and isinstance(col, int)) else "unknown"
        safe_sheet_name = secure_filename(sheet_name)
        image_filename = f"{safe_sheet_name}_img_{coord}_{idx}.png"
        save_path = self.media_dir / image_filename

        item = {"coord": coord, "filename": str(image_filename), "type": "image"}

        try:
            raw_data = self._get_image_raw_data(img, coord, sheet_name)
            if raw_data is None:
                return None

            with Image.open(io.BytesIO(raw_data)) as pillow_img:
                if pillow_img.mode in ("RGBA", "P"):
                    pillow_img = pillow_img.convert("RGB")
                pillow_img.save(save_path, "PNG")
            return item
        except Exception as e:
            logger.warning(f"Failed to extract image at {coord} on sheet {sheet_name}: {e}")
            item["filename"] = None
            item["error"] = "unidentified_format"
            return item

    def _extract_media(self, ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """
        ワークシートから埋め込み画像（図、グラフ、写真）を抽出し保存する。
        """
        media_info = []
        if not hasattr(ws, "_images"):
            return media_info

        for idx, img in enumerate(ws._images):
            item = self._process_single_image(img, idx, sheet_name)
            if item:
                media_info.append(item)

        return media_info

    def _get_merged_cells_map(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> Dict[Tuple[int, int], Union[str, Dict[str, int]]]:
        """
        シート内の結合セル情報をマップ化する。

        Returns:
            Dict: 左上セルの座標(r, c)をキーとし、スパン情報を値に持つ辞書。
        """
        merged_map = {}
        for m_range in ws.merged_cells.ranges:
            for r, c in m_range.cells:
                if r == m_range.min_row and c == m_range.min_col:
                    merged_map[(r, c)] = {
                        "colspan": m_range.max_col - m_range.min_col + 1,
                        "rowspan": m_range.max_row - m_range.min_row + 1,
                    }
                else:
                    merged_map[(r, c)] = "skip"
        return merged_map

    def _cell_to_html_td(
        self,
        cell: openpyxl.cell.Cell,
        span_info: Union[str, Dict],
        style_str: str,
        unit: Optional[str],
        formula: Optional[str] = None,
    ) -> str:
        """
        単一のセルを、詳細属性付きのHTML <td> タグに変換する。
        """
        val = cell.value
        val_str = (
            val.isoformat()
            if isinstance(val, (date, datetime))
            else str(clean_kami_text(val))
            if val is not None
            else ""
        )

        attrs = [f'data-coord="{cell.coordinate}"']
        if isinstance(span_info, dict):
            if span_info.get("colspan", 1) > 1:
                attrs.append(f'colspan="{span_info["colspan"]}"')
            if span_info.get("rowspan", 1) > 1:
                attrs.append(f'rowspan="{span_info["rowspan"]}"')

        if style_str:
            attrs.append(f'style="{style_str}"')

        if formula and str(formula).startswith("="):
            attrs.append(f'data-formula="{html.escape(str(formula))}"')

        if unit:
            attrs.append(f'data-unit="{html.escape(unit)}"')

        attr_str = " " + " ".join(attrs) if attrs else ""
        safe_val = html.escape(val_str).replace("\n", "<br>")
        return f"<td{attr_str}>{safe_val}</td>"

    def is_simple_table(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """シートが単純な表形式（結合なし、1行目見出し）かどうかを判定する。"""
        if ws.merged_cells.ranges:
            return False
        if ws.max_row < 2 or ws.max_column < 1:
            return False

        # 1行目に少なくとも2つの非空セルがあるか確認（早めに終了）
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        header_count = 0
        for val in first_row:
            if val is not None:
                header_count += 1
                if header_count >= 2:
                    return True
        return False

    def _extract_row_dict(self, row: Tuple[Any, ...], headers: List[str]) -> Dict[str, Any]:
        """1行のデータを辞書形式に変換する（日付処理・None除外を含む）。"""
        row_dict = {}
        for i, v in enumerate(row):
            if v is not None:
                row_dict[headers[i]] = v.isoformat() if isinstance(v, (date, datetime)) else v
        return row_dict

    def extract_simple_table(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> List[Dict[str, Any]]:
        """単純な表形式のシートからデータを高速に抽出する。"""
        rows_gen = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_gen)
        except StopIteration:
            return []

        headers = [str(v or f"Column{i + 1}") for i, v in enumerate(header_row)]
        data = []
        for row in rows_gen:
            row_dict = self._extract_row_dict(row, headers)
            if row_dict:
                data.append(row_dict)
        return data

    def _get_bounding_box(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        ws_formula: Optional[openpyxl.worksheet.worksheet.Worksheet] = None,
    ) -> Tuple[int, int, int, int]:
        """
        データまたは書式が存在する実質的な範囲（最小行、最大行、最小列、最大列）を特定する。
        """
        max_r, max_c = 0, 0

        # データがあるセルの範囲を取得
        # ⚡ Performance: Use iter_rows(values_only=True) to efficiently scan for data
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_has_data = False
            for c_idx, value in enumerate(row, 1):
                if value is not None:
                    row_has_data = True
                    if c_idx > max_c:
                        max_c = c_idx
            if row_has_data:
                max_r = max(max_r, r_idx)

        # フォームラがある範囲も含める
        if ws_formula:
            for r_idx, row in enumerate(ws_formula.iter_rows(values_only=True), 1):
                row_has_data = False
                for c_idx, value in enumerate(row, 1):
                    if value is not None and str(value).startswith("="):
                        row_has_data = True
                        if c_idx > max_c:
                            max_c = c_idx
                if row_has_data:
                    max_r = max(max_r, r_idx)

        # 結合セルや画像がある範囲も含める
        for m_range in ws.merged_cells.ranges:
            max_r = max(max_r, m_range.max_row)
            max_c = max(max_c, m_range.max_col)

        if hasattr(ws, "_images"):
            for img in ws._images:
                try:
                    row, col = self._parse_image_anchor(img.anchor)
                except Exception as e:
                    logger.warning(f"Failed to parse coordinate anchor '{img.anchor}' in bounding box calculation: {e}")
                    continue

                # 🔒 Robustness: Ensure we are comparing actual integers, not mocks
                if isinstance(row, int):
                    max_r = max(max_r, row)
                if isinstance(col, int):
                    max_c = max(max_c, col)

        return 1, max_r, 1, max_c

    def _generate_metadata_and_html(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        ws_formula: Optional[openpyxl.worksheet.worksheet.Worksheet] = None,
        merged_map: Optional[Dict] = None,
    ) -> Tuple[str, List[Dict[str, Any]]]:
        """詳細メタデータとHTMLテーブルを同時に生成する。"""
        if merged_map is None:
            merged_map = self._get_merged_cells_map(ws)

        min_r, max_r, min_c, max_c = self._get_bounding_box(ws, ws_formula=ws_formula)
        cell_metadata = []
        html_rows = ["<table border='1' style=\"border-collapse: collapse; min-width: 100%;\">"]

        # iter_rowsを使用してセルへのアクセスを高速化
        rows_gen = ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c)
        rows_f_gen = (
            ws_formula.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c) if ws_formula else None
        )

        for r_idx, row in enumerate(rows_gen, min_r):
            row_f = next(rows_f_gen) if rows_f_gen else None
            row_html = ["  <tr>"]

            current_row_html = []
            for c_idx, (cell, cell_f) in enumerate(zip(row, row_f if row_f else row), min_c):
                span = merged_map.get((r_idx, c_idx))
                if span == "skip":
                    continue

                formula = cell_f.value if row_f else None

                # ⚡ Performance: Cache style-related information to avoid redundant calculations
                s_id = cell.style_id
                if s_id not in self._style_cache:
                    self._style_cache[s_id] = (
                        self._get_cell_style_string(cell),
                        self._get_unit_info(cell),
                        self._get_border_info(cell),
                        bool(cell.font.b if cell.font else False),
                    )
                style_str, unit, borders, bold = self._style_cache[s_id]

                # メタデータの構築
                cell_info = {
                    "coord": cell.coordinate,
                    "row": r_idx,
                    "col": c_idx,
                    "value": str(clean_kami_text(cell.value)) if cell.value is not None else None,
                    "formula": formula if str(formula).startswith("=") else None,
                    "unit": unit,
                    "style": {
                        "borders": borders,
                        "bold": bold,
                    },
                }
                if isinstance(span, dict):
                    cell_info.update(span)
                cell_metadata.append(cell_info)

                # HTMLテーブル行の構築
                td_html = self._cell_to_html_td(cell, span, style_str, unit, formula=formula)
                current_row_html.append(td_html)

            # 結合セルなどの情報を考慮し、bounding box内の全行を出力
            row_html.extend(current_row_html)
            row_html.append("  </tr>")
            html_rows.append("".join(row_html))

        html_rows.append("</table>")
        return "\n".join(html_rows), cell_metadata

    def extract(self, excel_path: Path, include_logic: bool = False) -> Dict[str, Any]:
        """
        Excelファイルを解析し、詳細な構造、スタイル、ロジック、メディアを抽出する。

        Args:
            excel_path: 解析対象のエクセルファイルのパス。
            include_logic: 計算式(formula)の抽出を有効にするかどうか。

        Returns:
            Dict: 全シートの解析データを含む辞書。
        """
        self._style_cache = {}  # Reset cache for each new workbook
        # 値の抽出用に data_only=True でロード
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        wb_formula = openpyxl.load_workbook(excel_path, data_only=False) if include_logic else None

        full_map = {"sheets": {}}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_f = wb_formula[sheet_name] if wb_formula else None

            media_info = self._extract_media(ws, sheet_name)
            media_map = {}
            for m in media_info:
                coord = m.get("coord", "unknown")
                media_map.setdefault(coord, []).append(m)

            # 詳細メタデータの生成
            merged_map = self._get_merged_cells_map(ws)
            html_table, cell_metadata = self._generate_metadata_and_html(ws, ws_formula=ws_f, merged_map=merged_map)

            full_map["sheets"][sheet_name] = {
                "html": html_table,
                "cells": cell_metadata,
                "media": media_info,
                "media_map": media_map,
                "is_simple": self.is_simple_table(ws),
            }
            if full_map["sheets"][sheet_name]["is_simple"]:
                full_map["sheets"][sheet_name]["structured_data"] = self.extract_simple_table(ws)

        return full_map

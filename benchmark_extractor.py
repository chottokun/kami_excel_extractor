import time
import openpyxl
from pathlib import Path
from kami_excel_extractor.extractor import MetadataExtractor

def create_large_excel(path: Path, rows=500, cols=50):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LargeSheet"

    # Add some styles
    from openpyxl.styles import Font, Fill, Border, Side, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFEE11", end_color="FFEE11", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = f"Data {r}-{c}"
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = border
            if c % 5 == 0:
                cell.number_format = '#,##0'

    wb.save(path)

def run_benchmark():
    test_file = Path("benchmark_test.xlsx")
    if not test_file.exists():
        print("Creating test file...")
        create_large_excel(test_file, rows=1000, cols=50) # 50,000 cells

    extractor = MetadataExtractor(output_dir="benchmark_out")

    start_time = time.time()
    result = extractor.extract(test_file)
    end_time = time.time()

    print(f"Extraction took: {end_time - start_time:.4f} seconds")

    # Run again to see if there's any variation
    start_time = time.time()
    result = extractor.extract(test_file)
    end_time = time.time()
    print(f"Extraction (second run) took: {end_time - start_time:.4f} seconds")

if __name__ == "__main__":
    run_benchmark()
